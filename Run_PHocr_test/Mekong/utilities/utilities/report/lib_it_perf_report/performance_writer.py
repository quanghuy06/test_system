from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import Font, Border, Side


class PerformanceReportWriter:

    def __init__(self, workbook, change, base_delta, performance_data):
        self.workbook = workbook
        self.performance_data = performance_data
        self.delta = base_delta
        self.ws_pef = self.workbook.create_sheet("C{}-Pef".format(change))
        self.total_current_time = 0
        self.total_previous_time = 0
        for test_case, current_time, previous_time in self.performance_data:
            if current_time > 0:
                self.total_current_time += current_time
            if previous_time > 0:
                self.total_previous_time += previous_time

    def set_style(self):
        """
        Set style for performance report sheet
        Returns
        -------

        """
        side = Side(style='thin', color="000000")
        border = Border(left=side, top=side, right=side, bottom=side)
        self.ws_pef.column_dimensions['A'].width = 80
        self.ws_pef.column_dimensions['D'].width = 14
        self.ws_pef['A1'].border = border
        self.ws_pef['A1'].fill = PatternFill("solid", fgColor="95b3d7")
        self.ws_pef['A1'].font = Font(name='Arial', size=10, bold=True)
        self.ws_pef['A1'].alignment = Alignment(
            horizontal='right',
            vertical='center',
            wrap_text=True
        )

        self.ws_pef['B1'].fill = PatternFill("solid", fgColor="95b3d7")
        self.ws_pef['B1'].border = border
        self.ws_pef['B1'].font = Font(name='Arial', size=10)

        self.ws_pef['C1'].fill = PatternFill("solid", fgColor="95b3d7")
        self.ws_pef['C1'].border = border
        self.ws_pef['C1'].font = Font(name='Arial', size=10)

        if self.total_current_time <= 0:
            self.ws_pef['B1'].alignment = Alignment(horizontal='right')

        if self.total_previous_time <= 0:
            self.ws_pef['C1'].alignment = Alignment(horizontal='right')

        for cell in ['A2', 'B2', 'C2', 'D2']:
            self.ws_pef[cell].border = border
            self.ws_pef[cell].fill = PatternFill("solid", fgColor="ffff00")
            self.ws_pef[cell].font = Font(name='Arial', size=10)
            self.ws_pef[cell].alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

        for index, (test_case, dt_time, pre_data_time) in enumerate(self.performance_data):
            self.ws_pef['A{}'.format(index + 3)].border = border
            self.ws_pef['A{}'.format(index + 3)].fill = PatternFill("solid", fgColor="ebf1de")
            self.ws_pef['B{}'.format(index + 3)].border = border
            self.ws_pef['B{}'.format(index + 3)].fill = PatternFill("solid", fgColor="e4dfec")
            self.ws_pef['C{}'.format(index + 3)].border = border
            self.ws_pef['C{}'.format(index + 3)].fill = PatternFill("solid", fgColor="e4dfec")
            self.ws_pef['D{}'.format(index + 3)].border = border
            self.ws_pef['A{}'.format(index + 3)].font = Font(name='Arial', size=10)
            self.ws_pef['B{}'.format(index + 3)].font = Font(name='Arial', size=10)
            self.ws_pef['C{}'.format(index + 3)].font = Font(name='Arial', size=10)
            self.ws_pef['D{}'.format(index + 3)].font = Font(name='Arial', size=10)

            self.ws_pef['B{}'.format(index + 3)].number_format = '#,##0.00'
            self.ws_pef['C{}'.format(index + 3)].number_format = '#,##0.00'
            self.ws_pef['D{}'.format(index + 3)].number_format = '#,##0.00'

            if dt_time <= 0:
                self.ws_pef['B{}'.format(index + 3)].alignment = Alignment(horizontal='right')

            if pre_data_time <= 0:
                self.ws_pef['C{}'.format(index + 3)].alignment = Alignment(horizontal='right')

            if dt_time > 0 and pre_data_time > 0:
                self.ws_pef['D{}'.format(index + 3)].border = border
                if dt_time - pre_data_time > 0:
                    red_back_ground = PatternFill("solid", fgColor="ffc7ce")
                    self.ws_pef['D{}'.format(index + 3)].fill = red_back_ground
                if dt_time - pre_data_time < 0:
                    green_back_ground = PatternFill("solid", fgColor="c6efce")
                    self.ws_pef['D{}'.format(index + 3)].fill = green_back_ground
            else:
                self.ws_pef['D{}'.format(index + 3)].alignment = Alignment(horizontal='right')

    def write_data(self):
        """
        Writing data for performance report
        Returns
        -------

        """
        self.ws_pef['A1'] = 'Total time:'
        self.ws_pef['A2'] = 'Test name'
        self.ws_pef['B2'] = 'Current'
        self.ws_pef['C2'] = 'D{}'.format(self.delta)
        self.ws_pef['D2'] = 'Variance(C-D)'

        if self.total_current_time > 0:
            self.ws_pef['B1'] = self.total_current_time
        else:
            self.ws_pef['B1'] = '-'

        if self.total_previous_time > 0:
            self.ws_pef['C1'] = self.total_previous_time
        else:
            self.ws_pef['C1'] = '-'

        for index, (test_case, dt_time, pre_data_time) in enumerate(self.performance_data):
            self.ws_pef['A{}'.format(index + 3)] = test_case

            if dt_time > 0:
                self.ws_pef['B{}'.format(index + 3)] = dt_time
            else:
                self.ws_pef['B{}'.format(index + 3)] = '-'

            if pre_data_time > 0:
                self.ws_pef['C{}'.format(index + 3)] = pre_data_time
            else:
                self.ws_pef['C{}'.format(index + 3)] = '-'

            if dt_time > 0 and pre_data_time > 0:
                self.ws_pef['D{}'.format(index + 3)] = '=B{row} - C{row}'.format(row=index + 3)
            else:
                self.ws_pef['D{}'.format(index + 3)] = '-'

    def write_sheet(self):
        """
        Write sheet
        Returns
        -------

        """
        self.set_style()
        self.write_data()
