from openpyxl import load_workbook
import re


class PerfTable(object):
    START_ROW = 2
    TEST_CASE = 0
    TIME = 1


class PerformanceReportReader:
    def __init__(self, per_report):
        self._per_report = per_report
        self.workbook = load_workbook(self._per_report)
        self._delta = None
        self._change_number = None
        self._perf_sheet_name = None
        self._perf_sheet = None
        self._summary_sheet = None
        self._perf_report_data = None
        self._summary_report_data = None

    @property
    def perf_sheet(self):
        """
        Returns
        -------
        str
            performance sheet
        """
        if self._perf_sheet is None:
            self._perf_sheet = self.workbook.get_sheet_by_name(self.perf_sheet_name)
        return self._perf_sheet

    @property
    def delta(self):
        """

        Returns
        -------
        str
            data
        """
        """
        Get delta of performance sheet
        :return:
        """
        if self._delta is None:
            delta_tex = self.perf_sheet['C2'].value
            delta_pat = re.compile(r"D(\d+)")
            delta_mat = delta_pat.match(delta_tex)
            if delta_mat is not None:
                self._delta = delta_mat.group(1)
        return self._delta

    @property
    def perf_sheet_name(self):
        """

        Returns
        -------
        str
            performance sheet name
        """
        if self._perf_sheet_name is None:
            self._perf_sheet_name = "C{}-Pef".format(self.change_number)
        return self._perf_sheet_name

    @property
    def change_number(self):
        """

        Returns
        -------
        str
            change number
        """
        if self._change_number is None:
            per_sheet_pat = re.compile(r"C(\d+)-Pef")
            for sheet_name in self.workbook.sheetnames:
                per_sheet_mat = per_sheet_pat.match(sheet_name)
                if per_sheet_mat is not None:
                    self._change_number = per_sheet_mat.group(1)
        return self._change_number

    @property
    def perf_report_data(self):
        """
        Get data of performance sheet.
        Returns
        -------
        dict
            A dictionary with key is name of test case and value is time running of the
            test case
            For example:
            {
                "01_0033": 3.14,
                ...
            }
        """
        if self._perf_report_data is None:
            self._perf_report_data = dict()
            for curr_row, cell in enumerate(self.perf_sheet.rows):
                if curr_row >= PerfTable.START_ROW:
                    test_case = cell[PerfTable.TEST_CASE].value
                    time = cell[PerfTable.TIME].value
                    self._perf_report_data[test_case] = time
        return self._perf_report_data

    def total_time(self):
        """
        Returns
        -------
        int:
            total time running performance for performance sheet.

        """
        total = 0
        for key in self.perf_report_data:
            total += self.perf_report_data.get(key)
        return total

