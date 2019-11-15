from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import Font, Border, Side


class SummaryReportWriter:

    def __init__(self, workbook, summary_report_data):
        self.workbook = workbook
        self.sheet_summary = self.workbook.create_sheet("Summary")
        self.summary_report_data = summary_report_data

    def set_style(self):
        """
        Set style for summary report
        Returns
        -------

        """
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        row_dimensions = self.sheet_summary.row_dimensions[3]
        row_dimensions.height = 50
        self.sheet_summary.column_dimensions['D'].width = 18
        self.sheet_summary.column_dimensions['E'].width = 28
        self.sheet_summary.column_dimensions['F'].width = 20
        self.sheet_summary.column_dimensions['H'].width = 105

        self.sheet_summary['E2'].font = Font(name='Arial', size=10)
        self.sheet_summary['F2'].font = Font(name='Arial', size=10)

        for cell in ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3']:
            self.sheet_summary[cell].border = Border(left=bd, top=bd, right=bd, bottom=bd)
            self.sheet_summary[cell].fill = PatternFill("solid", fgColor="95b3d7")
            self.sheet_summary[cell].alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True)
            self.sheet_summary[cell].font = Font(name='Arial', size=10, bold=True)

        for row_index, row_value in enumerate(self.summary_report_data):
            self.sheet_summary['B{}'.format(row_index + 4)].border = border
            self.sheet_summary['C{}'.format(row_index + 4)].border = border
            self.sheet_summary['D{}'.format(row_index + 4)].border = border
            self.sheet_summary['E{}'.format(row_index + 4)].border = border
            self.sheet_summary['F{}'.format(row_index + 4)].border = border

            self.sheet_summary['B{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['C{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['D{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['E{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['F{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['G{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['H{}'.format(row_index + 4)].font = Font(name='Arial', size=10)
            self.sheet_summary['E{}'.format(row_index + 4)].number_format = '#,##0.00'
            if row_index > 0:
                delta = row_value.performance - self.summary_report_data[row_index - 1].performance
                if delta < 0:
                    green_bg = PatternFill("solid", fgColor="c6efce")
                    self.sheet_summary['F{}'.format(row_index + 4)].fill = green_bg
                if delta > 0:
                    red_bg = PatternFill("solid", fgColor="ffc7ce")
                    self.sheet_summary['F{}'.format(row_index + 4)].fill = red_bg
            self.sheet_summary['F{}'.format(row_index + 4)].number_format = '0.00%'
            self.sheet_summary['G{}'.format(row_index + 4)].border = border
            self.sheet_summary['H{}'.format(row_index + 4)].border = border

            if row_index < len(self.summary_report_data) - 10:
                self.sheet_summary.row_dimensions[row_index + 4].hidden = True

    def write_data(self):
        """
        Writing data for summary report
        Returns
        -------

        """
        self.sheet_summary['E2'] = 'Total character in 96 test cases:'
        self.sheet_summary['F2'] = 'Latest merged version:'

        self.sheet_summary['B3'] = 'Build'
        self.sheet_summary['C3'] = 'Version'
        self.sheet_summary['D3'] = 'Gerrit number'
        self.sheet_summary['E3'] = 'Performance'
        self.sheet_summary['F3'] = 'Improvement (compare with base version)'
        self.sheet_summary['G3'] = 'Base version'
        self.sheet_summary['H3'] = 'Note'

        for row_index, row_value in enumerate(self.summary_report_data):
            self.sheet_summary['B{}'.format(row_index + 4)] = row_value.build
            self.sheet_summary['C{}'.format(row_index + 4)] = row_value.version
            self.sheet_summary['D{}'.format(row_index + 4)] = row_value.gerrit_number
            self.sheet_summary['E{}'.format(row_index + 4)] = row_value.performance
            self.sheet_summary['F{}'.format(row_index + 4)] = row_value.improvement
            self.sheet_summary['G{}'.format(row_index + 4)] = row_value.base_version
            self.sheet_summary['H{}'.format(row_index + 4)] = row_value.note

    def write_sheet(self):
        """
        Writing summary report sheet
        Returns
        -------

        """
        self.set_style()
        self.write_data()
